#!/usr/bin/env python3
"""
从 xlsx 文件导入数据到 data.json
"""

import openpyxl
import json
import os
import re
import random
import string
import time
from datetime import datetime

DATA_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data.json')

# 类型映射
TYPE_MAP = {
    '📚 学习输入': 'learning',
    '📚学习输入': 'learning',
    '🏃🏻‍♀️ 运动锻炼': 'exercise',
    '🏃‍♀️ 运动锻炼': 'exercise',
    '🎵 兴趣爱好': 'hobby',
    '🌲 目标设定': 'goal',
    '🚗 其他事项': 'other',
    '学习输入': 'learning',
    '运动锻炼': 'exercise',
    '兴趣爱好': 'hobby',
    '目标设定': 'goal',
    '其他事项': 'other',
}

def generate_id():
    """生成唯一ID"""
    return str(int(time.time() * 1000)) + ''.join(random.choices(string.ascii_lowercase + string.digits, k=9))

def parse_date(date_str):
    """解析日期字符串为 YYYY-MM-DD 格式"""
    if not date_str:
        return None
    
    # 如果是 datetime 对象
    if isinstance(date_str, datetime):
        return date_str.strftime('%Y-%m-%d')
    
    date_str = str(date_str).strip()
    
    # 尝试多种格式
    patterns = [
        (r'(\d{4})年(\d{1,2})月(\d{1,2})日', lambda m: f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"),
        (r'(\d{4})-(\d{1,2})-(\d{1,2})', lambda m: f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"),
        (r'(\d{4})/(\d{1,2})/(\d{1,2})', lambda m: f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"),
    ]
    
    for pattern, formatter in patterns:
        match = re.search(pattern, date_str)
        if match:
            return formatter(match)
    
    return None

def parse_month(month_str, date_str=None):
    """解析月份"""
    if month_str and '月' in str(month_str):
        match = re.search(r'(\d{1,2})月', str(month_str))
        if match:
            return f"{int(match.group(1))}月"
    
    # 从日期中提取
    if date_str:
        parsed = parse_date(date_str)
        if parsed:
            month = int(parsed.split('-')[1])
            return f"{month}月"
    
    return None

def parse_week(week_str):
    """解析周"""
    if not week_str:
        return None
    week_str = str(week_str)
    if '第' in week_str and '周' in week_str:
        return week_str
    return None

def parse_type(type_str):
    """解析类型"""
    if not type_str:
        return 'learning'
    
    type_str = str(type_str).strip()
    
    # 直接匹配
    if type_str in TYPE_MAP:
        return TYPE_MAP[type_str]
    
    # 部分匹配
    for key, value in TYPE_MAP.items():
        if key in type_str or type_str in key:
            return value
    
    # 根据关键词判断
    if '学习' in type_str or '读书' in type_str or '英语' in type_str:
        return 'learning'
    elif '运动' in type_str or '跑步' in type_str or '锻炼' in type_str:
        return 'exercise'
    elif '兴趣' in type_str or '爱好' in type_str or '声乐' in type_str or '音乐' in type_str:
        return 'hobby'
    elif '其他' in type_str:
        return 'other'
    
    return 'learning'

def parse_completed(completed_str):
    """解析是否完成"""
    if not completed_str:
        return False
    
    completed_str = str(completed_str).strip()
    
    if completed_str in ['☑', '✅', '✓', '√', '1', 'True', 'true', 'yes', 'Yes', '是', '完成']:
        return True
    if completed_str in ['□', '☐', '0', 'False', 'false', 'no', 'No', '否', '未完成']:
        return False
    
    return False

def read_xlsx(file_path):
    """读取 xlsx 文件并转换为计划列表"""
    print(f"\n📂 正在读取: {file_path}")
    
    wb = openpyxl.load_workbook(file_path)
    plans = []
    
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"  📄 工作表: {sheet_name}, 共 {sheet.max_row} 行")
        
        # 跳过第1行(说明)和第2行(表头)
        for row_idx, row in enumerate(sheet.iter_rows(min_row=3, values_only=True), 3):
            # 跳过空行
            if not row[0]:
                continue
            
            title = str(row[0]).strip() if row[0] else None
            if not title:
                continue
            
            # 跳过目标设定行（通常是说明性文字）
            if '目标：' in title or '目标:' in title:
                continue
            
            date = parse_date(row[1])
            month = parse_month(row[2], row[1])
            week = parse_week(row[3])
            plan_type = parse_type(row[4])
            completed = parse_completed(row[5])
            summary = str(row[7]).strip() if row[7] else ''
            
            # 必须有日期才导入
            if not date:
                print(f"    ⚠️ 跳过(无日期): {title[:30]}...")
                continue
            
            plan = {
                'id': generate_id(),
                'title': title,
                'date': date,
                'month': month or '',
                'week': week or '',
                'type': plan_type,
                'completed': completed,
                'summary': summary
            }
            plans.append(plan)
            time.sleep(0.001)  # 确保 ID 唯一
    
    print(f"  ✅ 解析出 {len(plans)} 条有效计划")
    return plans

def load_existing_data():
    """加载现有数据"""
    try:
        with open(DATA_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    except:
        return []

def save_data(data):
    """保存数据"""
    with open(DATA_FILE, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def main():
    import sys
    
    # 获取所有 xlsx 文件
    xlsx_files = [f for f in os.listdir(os.path.dirname(os.path.abspath(__file__)) or '.') 
                  if f.endswith('.xlsx')]
    
    if not xlsx_files:
        print("❌ 没有找到 xlsx 文件")
        return
    
    print("=" * 50)
    print("📊 XLSX 数据导入工具")
    print("=" * 50)
    print("\n可用的 xlsx 文件:")
    for i, f in enumerate(xlsx_files, 1):
        print(f"  {i}. {f}")
    
    # 命令行参数
    if len(sys.argv) > 1:
        if sys.argv[1] == '--all':
            selected_files = xlsx_files
        else:
            selected_files = [f for f in sys.argv[1:] if f in xlsx_files]
    else:
        print("\n使用方法:")
        print("  python3 import_xlsx.py <文件名>     导入指定文件")
        print("  python3 import_xlsx.py --all        导入所有文件")
        return
    
    if not selected_files:
        print("❌ 没有有效的文件")
        return
    
    # 导入数据
    all_plans = []
    for file_name in selected_files:
        file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), file_name)
        plans = read_xlsx(file_path)
        all_plans.extend(plans)
    
    if not all_plans:
        print("\n❌ 没有解析出有效数据")
        return
    
    # 询问是否覆盖
    existing = load_existing_data()
    print(f"\n📋 现有数据: {len(existing)} 条")
    print(f"📥 待导入数据: {len(all_plans)} 条")
    
    choice = input("\n选择导入方式 (1=追加, 2=覆盖, 其他=取消): ").strip()
    
    if choice == '1':
        # 追加
        existing.extend(all_plans)
        save_data(existing)
        print(f"\n✅ 成功追加 {len(all_plans)} 条数据，现共 {len(existing)} 条")
    elif choice == '2':
        # 覆盖
        save_data(all_plans)
        print(f"\n✅ 成功覆盖，现共 {len(all_plans)} 条数据")
    else:
        print("\n❌ 已取消")

if __name__ == '__main__':
    main()

